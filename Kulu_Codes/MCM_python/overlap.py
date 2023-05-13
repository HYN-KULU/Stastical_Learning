import pandas as pd
import matplotlib.pyplot as plt  # 画图用
import numpy as np
import matplotlib
import xlrd
import openpyxl
file_location = "./Problem_C_Data_Wordle.xlsx"
# data是Excel里的数据
data = pd.read_excel(file_location)  # 默认读取第一个sheet
word_list = []
count_list = []
for i in range(1, 357):
    word_list.append(data.iloc[i, 3])
for i in range(0, 356):
    count = 0
    word = word_list[i]
    for i in range(0, len(word)-1):
        for j in range(i+1, len(word)):
            if word[i] == word[j]:
                count = count+1
    count_list.append(count)
print(count_list)
workbook = openpyxl.Workbook()  # workbook = openpyxl.Workbook(encoding='UTF-8')
# 获取活动工作表， 默认就是第一个工作表
worksheet = workbook.active
worksheet.title = "FJ命名"

# 新建工作表，尾追加创建，或者头部创建
# worksheet2 = workbook.create_sheet()    # 默认在尾部创建
worksheet2 = workbook.create_sheet(0)   # 在头部创建
worksheet2.title = "添加表名字"

# 写入第一列数据， 因为第一行已经有数据了，所以是 i + 2
for i in range(len(count_list)):
    worksheet.cell(i + 2, 1, count_list[i])
workbook.save(filename="overlap.xlsx")

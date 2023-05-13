import openpyxl
workbook = openpyxl.Workbook()  # workbook = openpyxl.Workbook(encoding='UTF-8')
# 获取活动工作表， 默认就是第一个工作表
worksheet = workbook.active
worksheet.title = "FJ命名"

# 新建工作表，尾追加创建，或者头部创建
# worksheet2 = workbook.create_sheet()    # 默认在尾部创建
worksheet2 = workbook.create_sheet(0)   # 在头部创建
worksheet2.title = "添加表名字"

# 写入第一列数据， 因为第一行已经有数据了，所以是 i + 2
for i in range(len(Province)):
    worksheet.cell(i + 2, 1, Province[i])


workbook.save(filename="09newTest.xlsx")
